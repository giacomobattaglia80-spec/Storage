#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""KPI Scorecard with RAG traffic lights (target vs actual)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import IconSetRule, FormulaRule

OUT = "/home/user/Storage/Hennecke/candidacy/KPI_Scorecard_HenneckeOMS.xlsx"
BLUE="1F3A5F"; GREEN="2E7D4F"; RED="C8102E"; GOLD="B8860B"; YEL="FFF2CC"
wb=Workbook(); ws=wb.active; ws.title="Scorecard"
ws.sheet_view.showGridLines=False
white=Font(color="FFFFFF",bold=True,name="Calibri",size=11)
bold=Font(bold=True,name="Calibri",size=11)
ital=Font(italic=True,color="666666",size=9.5,name="Calibri")
title_f=Font(bold=True,size=15,color=BLUE,name="Calibri")
hdr=PatternFill("solid",fgColor=BLUE)
inp=PatternFill("solid",fgColor=YEL)
thin=Side(style="thin",color="C9D2DD"); box=Border(left=thin,right=thin,top=thin,bottom=thin)
GREENF=PatternFill("solid",fgColor="CDECD6"); AMBERF=PatternFill("solid",fgColor="FCE4B6"); REDF=PatternFill("solid",fgColor="F8CBCB")

widths={"A":2,"B":34,"C":50,"D":8,"E":10,"F":12,"G":12,"H":9}
for c,w in widths.items(): ws.column_dimensions[c].width=w

def put(cell,val,*,f=None,fill=None,nf=None,align="left",border=True,wrap=False):
    c=ws[cell]; c.value=val
    if f:c.font=f
    if fill:c.fill=fill
    if nf:c.number_format=nf
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if border:c.border=box
    return c

def band(rng,text):
    for row in ws[rng]:
        for c in row: c.fill=hdr; c.font=white; c.alignment=Alignment(vertical="center")
    ws[rng.split(":")[0]].value=text; ws.merge_cells(rng)

ws.merge_cells("B1:H1"); put("B1","KPI Scorecard — HENNECKE-OMS  ·  crescita · redditività · cassa",f=title_f,border=False)
ws.merge_cells("B2:H2"); put("B2","Compila le celle GIALLE (Target e Attuale): il semaforo si aggiorna da solo. ▲ = meglio se alto · ▼ = meglio se basso. % in formato percentuale.",f=ital,border=False)

# header row
HR=4
for col,txt,al in [("B","KPI","left"),("C","Definizione / formula","left"),("D","Verso","center"),
                   ("E","Unità","center"),("F","Target","center"),("G","Attuale","center"),("H","Status","center")]:
    put(f"{col}{HR}",txt,f=white,fill=hdr,align=al)

PCT='0.0%'; NUM='#,##0.0'; INT='#,##0'

# (KPI, definizione, verso, unità, target, attuale, numfmt)
rows=[
 ("CRESCITA (Growth)",),
 ("Order intake","Nuovi ordini acquisiti nel periodo","▲","€m",50,44,NUM),
 ("Book-to-bill","Bookings ÷ ricavi fatturati","▲","x",1.05,0.95,NUM),
 ("Backlog — copertura","Backlog ÷ ricavi medi mensili","▲","mesi",9,7,NUM),
 ("Aftermarket share","Ricavi service/ricambi ÷ ricavi totali","▲","%",0.30,0.22,PCT),
 ("Win rate offerte","Valore vinto ÷ valore quotato","▲","%",0.30,0.25,PCT),
 ("REDDITIVITÀ (Profitability)",),
 ("EBITDA margin (norm.)","EBITDA ÷ ricavi — vista normalizzata 2-yr","▲","%",0.07,0.055,PCT),
 ("Margine-a-finire medio","(Prezzo − EAC) ÷ prezzo, medio commesse","▲","%",0.08,0.05,PCT),
 ("Material cost ratio","Acquisti ÷ valore della produzione","▼","%",0.35,0.384,PCT),
 ("Ricavi per FTE","Ricavi ÷ organico medio","▲","€k",290,266,INT),
 ("Costi in garanzia","Costi garanzia/claim ÷ ricavi","▼","%",0.015,0.020,PCT),
 ("CASSA (Cash)",),
 ("DIO magazzino","Rimanenze ÷ costo del venduto × 365","▼","gg",60,66,INT),
 ("DSO clienti","Crediti v/clienti ÷ ricavi × 365","▼","gg",10,6,INT),
 ("DPO fornitori","Debiti v/fornitori ÷ acquisti × 365","▲","gg",90,87,INT),
 ("Cash Conversion Cycle","DSO + DIO − DPO","▼","gg",10,-15,INT),
 ("Acconti ÷ WIP","Acconti clienti ÷ lavori in corso","▲","x",1.0,3.2,NUM),
 ("FCF conversion","Free cash flow ÷ EBITDA","▲","%",0.60,0.50,PCT),
]
r=HR+1
data_rows=[]
for row in rows:
    if len(row)==1:
        ws.merge_cells(f"B{r}:H{r}")
        for col in "BCDEFGH": ws[f"{col}{r}"].fill=PatternFill("solid",fgColor="E6EBF2"); ws[f"{col}{r}"].border=box
        cc=ws[f"B{r}"]; cc.value=row[0]; cc.font=Font(bold=True,color=BLUE,size=11); cc.alignment=Alignment(vertical="center")
        r+=1; continue
    kpi,defi,verso,unit,tgt,act,nf=row
    put(f"B{r}",kpi,f=bold)
    put(f"C{r}",defi,f=ital)
    put(f"D{r}",verso,align="center")
    put(f"E{r}",unit,align="center")
    put(f"F{r}",tgt,fill=inp,nf=nf,align="center")
    put(f"G{r}",act,fill=inp,nf=nf,align="center")
    # status score: 3 green / 2 amber / 1 red
    put(f"H{r}", f'=IF(G{r}="","",IF(D{r}="▲",IF(G{r}>=F{r},3,IF(G{r}>=F{r}*0.9,2,1)),IF(G{r}<=F{r},3,IF(G{r}<=F{r}*1.1,2,1))))',
        align="center")
    data_rows.append(r)
    r+=1
last=r-1

# traffic-light icon set on Status (H), hide the number
rng=f"H{HR+1}:H{last}"
icon=IconSetRule('3TrafficLights1','num',[0,2,3],showValue=False)
ws.conditional_formatting.add(rng,icon)
# colour the Attuale (G) cell by status
for rr in data_rows:
    ws.conditional_formatting.add(f"G{rr}",FormulaRule(formula=[f"$H{rr}=3"],fill=GREENF))
    ws.conditional_formatting.add(f"G{rr}",FormulaRule(formula=[f"$H{rr}=2"],fill=AMBERF))
    ws.conditional_formatting.add(f"G{rr}",FormulaRule(formula=[f"$H{rr}=1"],fill=REDF))

# legend & notes
ln=last+2
ws.merge_cells(f"B{ln}:H{ln}"); ws[f"B{ln}"]="Legenda semaforo: ● verde = a/oltre target · ● ambra = entro 10% dal target · ● rosso = sotto soglia."; ws[f"B{ln}"].font=ital
notes=[
 "I valori 'Attuale' pre-caricati sono in parte dal bilancio 2024 (es. EBITDA norm. 5,6% · material ratio 38,4% · DIO 66 · DSO 6 · DPO 87) e in parte esempi da popolare con i dati gestionali.",
 "Soglia ambra = ±10% dal target (modificabile nelle formule colonna H). Verso ▲ = meglio se alto, ▼ = meglio se basso.",
 "Collegabile al simulatore di commessa (margine-a-finire, EAC, WIP, acconti) per alimentare automaticamente i KPI operativi.",
]
for i,t in enumerate(notes):
    ws.merge_cells(f"B{ln+1+i}:H{ln+1+i}"); ws[f"B{ln+1+i}"]=t; ws[f"B{ln+1+i}"].font=ital

ws.freeze_panes=f"B{HR+1}"
wb.save(OUT); print("Saved:",OUT,"| KPI rows:",len(data_rows))

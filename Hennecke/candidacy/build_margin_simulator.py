#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Interactive project-margin simulator (Excel, live formulas)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter

OUT = "/home/user/Storage/Hennecke/candidacy/Simulatore_Margine_Commessa.xlsx"
wb = Workbook(); ws = wb.active; ws.title = "Margine Commessa"

BLUE = "1F3A5F"; YEL = "FFF2CC"; YELB = "E6B800"; GREYH = "EDF1F6"
GREEN = "2E7D4F"; RED = "C8102E"
white = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
hdr = PatternFill("solid", fgColor=BLUE)
inputfill = PatternFill("solid", fgColor=YEL)
greyfill = PatternFill("solid", fgColor=GREYH)
thin = Side(style="thin", color="C9D2DD")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
bold = Font(bold=True, name="Calibri", size=11)
ital = Font(italic=True, color="666666", size=9.5, name="Calibri")
title_f = Font(bold=True, size=15, color=BLUE, name="Calibri")

EUR = '#,##0'; PCT = '0.0%'

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 13
ws.column_dimensions["E"].width = 15
ws.column_dimensions["F"].width = 42

def put(cell, val, *, f=None, fill=None, nf=None, align=None, border=False, wrap=False):
    c = ws[cell]; c.value = val
    if f: c.font = f
    if fill: c.fill = fill
    if nf: c.number_format = nf
    if align: c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    else: c.alignment = Alignment(vertical="center", wrap_text=wrap)
    if border: c.border = box
    return c

def band(rng, text):
    for row in ws[rng]:
        for c in row:
            c.fill = hdr; c.font = white; c.alignment = Alignment(vertical="center")
    ws[rng.split(":")[0]].value = text
    ws.merge_cells(rng)

# ---- title ----
ws.merge_cells("B1:F1"); put("B1", "Simulatore Margine di Commessa — project margin", f=title_f)
ws.merge_cells("B2:F2"); put("B2", "Modifica SOLO le celle gialle (INPUT): tutto il resto si aggiorna in automatico. Importi in € migliaia.", f=ital)

# ---- INPUT ----
band("B4:D4", "INPUT  (celle gialle)")
inputs = [
    ("B5", "Prezzo commessa (ricavi totali)", "C5", 10000),
    ("B6", "Budget costi iniziale (offerta)", "C6", 8000),
    ("B7", "Costi sostenuti a oggi", "C7", 4300),
    ("B8", "Costi a finire — ETC (stima)", "C8", 4300),
    ("B9", "Fatturato a oggi (SAL emessi)", "C9", 3000),
    ("B10", "Acconti incassati (cassa)", "C10", 3000),
]
for bcell, label, ccell, val in inputs:
    put(bcell, label, border=True)
    put(ccell, val, f=bold, fill=inputfill, nf=EUR, align="right", border=True)
    put("D" + bcell[1:], "€ migliaia", f=ital, border=True)

# ---- RESULTS ----
band("B12:D12", "RISULTATI (calcolati in automatico)")
res = [
    ("B13", "EAC — costi totali stimati a fine commessa", "C13", "=C7+C8", EUR, "Estimate At Completion = costi a oggi + ETC"),
    ("B14", "Scostamento EAC vs budget (€)", "C14", "=C13-C6", EUR, "positivo = sforamento costi (overrun)"),
    ("B15", "Scostamento EAC vs budget (%)", "C15", "=(C13-C6)/C6", PCT, ""),
    ("B16", "% completamento (cost-to-cost)", "C16", "=C7/C13", PCT, "= costi a oggi ÷ EAC (non ÷ budget!)"),
    ("B17", "Margine di OFFERTA (€)", "C17", "=C5-C6", EUR, "quello promesso quando hai vinto la gara"),
    ("B18", "Margine di OFFERTA (%)", "C18", "=(C5-C6)/C5", PCT, ""),
    ("B19", "Margine-A-FINIRE (€)", "C19", "=C5-C13", EUR, "margine atteso sull'intera commessa, ricalcolato"),
    ("B20", "Margine-A-FINIRE (%)", "C20", "=(C5-C13)/C5", PCT, "< offerta = erosione di margine"),
    ("B21", "Erosione margine vs offerta (pp)", "C21", "=C20-C18", PCT, "differenza in punti percentuali"),
    ("B22", "Ricavi riconosciuti a oggi (POC)", "C22", "=C5*C16", EUR, "= prezzo × % completamento"),
    ("B23", "Costi riconosciuti a oggi", "C23", "=C7", EUR, ""),
    ("B24", "Margine riconosciuto a oggi (€)", "C24", "=C22-C23", EUR, ""),
    ("B25", "Contract ASSET — WIP (attivo)", "C25", "=MAX(C22-C9,0)", EUR, "ricavi maturati ma non ancora fatturati"),
    ("B26", "Contract LIABILITY (passivo)", "C26", "=MAX(C9-C22,0)", EUR, "fatturato/acconti in eccesso sui ricavi maturati"),
    ("B27", "Perdita attesa da ACCANTONARE (onerous)", "C27", "=IF(C13>C5,C13-C5,0)", EUR, "se EAC > prezzo: accantona SUBITO l'intera perdita"),
]
for bcell, label, ccell, formula, nf, note in res:
    r = bcell[1:]
    is_key = label.split()[0] in ("Margine-A-FINIRE", "EAC", "Perdita")
    put(bcell, label, f=(bold if is_key else None), border=True)
    put(ccell, formula, f=bold, nf=nf, align="right", border=True, fill=(greyfill if is_key else None))
    put("D" + r, note, f=ital, border=True, wrap=False)

# status flag
put("B28", "STATO COMMESSA", f=bold, border=True)
put("C28", '=IF(C19<0,"IN PERDITA — accantona",IF(C20<C18,"MARGINE IN EROSIONE","IN LINEA / OK"))',
    f=bold, align="center", border=True)
ws.merge_cells("D28:F28"); put("D28", "semaforo: rosso = perdita · ambra = erosione vs offerta · verde = in linea", f=ital, border=True)

# ---- SENSITIVITY ----
band("B30:E30", "SENSITIVITY — come si muove il margine al variare dei COSTI A FINIRE (ETC)")
put("B31", "ETC (costi a finire)", f=bold, fill=greyfill, align="center", border=True)
put("C31", "EAC", f=bold, fill=greyfill, align="center", border=True)
put("D31", "Margine-a-finire %", f=bold, fill=greyfill, align="center", border=True)
put("E31", "Δ vs offerta (pp)", f=bold, fill=greyfill, align="center", border=True)
etc_vals = [3000, 3500, 4000, 4300, 4800, 5300, 5800, 6300]
r0 = 32
for i, etc in enumerate(etc_vals):
    r = r0 + i
    put(f"B{r}", etc, nf=EUR, align="right", border=True)
    put(f"C{r}", f"=$C$7+B{r}", nf=EUR, align="right", border=True)
    put(f"D{r}", f"=($C$5-C{r})/$C$5", nf=PCT, align="right", border=True)
    put(f"E{r}", f"=D{r}-$C$18", nf=PCT, align="right", border=True)
last = r0 + len(etc_vals) - 1

# ---- conditional formatting ----
# margin-at-completion % (C20): green if >= offerta, red if <
ws.conditional_formatting.add("C20", FormulaRule(formula=["$C$20>=$C$18"], font=Font(color=GREEN, bold=True)))
ws.conditional_formatting.add("C20", FormulaRule(formula=["$C$20<$C$18"], font=Font(color=RED, bold=True)))
# margin-at-completion € (C19) red if negative
ws.conditional_formatting.add("C19", FormulaRule(formula=["$C$19<0"], font=Font(color=RED, bold=True)))
# provision (C27) red fill if > 0
ws.conditional_formatting.add("C27", FormulaRule(formula=["$C$27>0"], fill=PatternFill("solid", fgColor="F8CBCB"), font=Font(color=RED, bold=True)))
# status flag colors
ws.conditional_formatting.add("C28", FormulaRule(formula=['$C$19<0'], fill=PatternFill("solid", fgColor="F8CBCB"), font=Font(color=RED, bold=True)))
ws.conditional_formatting.add("C28", FormulaRule(formula=['AND($C$19>=0,$C$20<$C$18)'], fill=PatternFill("solid", fgColor="FCE4B6"), font=Font(color="9A6700", bold=True)))
ws.conditional_formatting.add("C28", FormulaRule(formula=['$C$20>=$C$18'], fill=PatternFill("solid", fgColor="CDECD6"), font=Font(color=GREEN, bold=True)))
# sensitivity colour scale on margin-at-completion %
ws.conditional_formatting.add(f"D{r0}:D{last}",
    ColorScaleRule(start_type="num", start_value=-0.1, start_color="F8696B",
                   mid_type="num", mid_value=0.1, mid_color="FFEB84",
                   end_type="num", end_value=0.25, end_color="63BE7B"))

# ---- notes block ----
nb = last + 2
ws.merge_cells(f"B{nb}:F{nb}"); put(f"B{nb}", "Come leggerlo (logica di controllo commessa):", f=bold)
notes = [
    "1) L'EAC è il numero chiave: alimenta sia la % di completamento (cost-to-cost) sia il margine-a-finire.",
    "2) Aumenta l'ETC (cella C8) e guarda C20: il margine-a-finire scende → vedi l'erosione in tempo reale.",
    "3) Se l'EAC supera il prezzo (C13>C5), C19 diventa negativo: scatta la perdita da accantonare SUBITO (onerous contract).",
    "4) Ricavi a bilancio = prezzo × % completamento (C22), non il fatturato: la differenza genera WIP/contract asset (C25) o contract liability (C26).",
    "5) Prova: porta C8 da 4.300 a 6.300 → EAC 10.600 > prezzo 10.000 → commessa in perdita.",
]
for i, t in enumerate(notes):
    ws.merge_cells(f"B{nb+1+i}:F{nb+1+i}"); put(f"B{nb+1+i}", t, f=ital)

ws.sheet_view.showGridLines = False
ws.freeze_panes = "B4"

# =====================================================================
# SHEET 2 — mini CONTO ECONOMICO & STATO PATRIMONIALE (linked)
# =====================================================================
ws2 = wb.create_sheet("CE & SP")
ws2.sheet_view.showGridLines = False
for col, w in {"A": 2, "B": 44, "C": 14, "D": 3, "E": 44, "F": 14}.items():
    ws2.column_dimensions[col].width = w
MC = "'Margine Commessa'!"

def c2(ref, val, *, f=None, fill=None, nf=None, align="left", border=True):
    c = ws2[ref]; c.value = val
    if f: c.font = f
    if fill: c.fill = fill
    if nf: c.number_format = nf
    c.alignment = Alignment(horizontal=align, vertical="center")
    if border: c.border = box
    return c

def band2(rng, text):
    for row in ws2[rng]:
        for cc in row:
            cc.fill = hdr; cc.font = white; cc.alignment = Alignment(vertical="center")
    ws2[rng.split(":")[0]].value = text
    ws2.merge_cells(rng)

ws2.merge_cells("B1:F1"); ws2["B1"] = "Mini Conto Economico & Stato Patrimoniale — collegati alla commessa"; ws2["B1"].font = title_f
ws2.merge_cells("B2:F2"); ws2["B2"] = "Si aggiornano in automatico al variare degli input del foglio 'Margine Commessa'. Importi in € migliaia."; ws2["B2"].font = ital

# ---------------- CONTO ECONOMICO (left) ----------------
band2("B4:C4", "CONTO ECONOMICO — a commessa, cumulato a oggi")
c2("B5", "A) Valore della produzione", f=bold)
c2("B6", "   Ricavi (SAL fatturati)"); c2("C6", f"={MC}C9", nf=EUR, align="right")
c2("B7", "   Variazione lavori in corso su ordinazione"); c2("C7", f"={MC}C22-{MC}C9", nf=EUR, align="right")
c2("B8", "   Totale valore della produzione", f=bold); c2("C8", "=C6+C7", f=bold, nf=EUR, align="right", fill=greyfill)
c2("B9", "B) Costi della produzione", f=bold)
c2("B10", "   Costi di commessa (materie/servizi/personale)"); c2("C10", f"={MC}C7", nf=EUR, align="right")
c2("B11", "   Accantonamento perdita attesa su commessa"); c2("C11", f"=IF({MC}C19<0,{MC}C24-{MC}C19,0)", nf=EUR, align="right")
c2("B12", "   Totale costi della produzione", f=bold); c2("C12", "=C10+C11", f=bold, nf=EUR, align="right", fill=greyfill)
c2("B13", "RISULTATO OPERATIVO (EBIT)", f=bold); c2("C13", "=C8-C12", f=bold, nf=EUR, align="right", fill=greyfill)
c2("B14", "   EBIT margin %"); c2("C14", "=IF(C8=0,0,C13/C8)", nf=PCT, align="right")
c2("B15", "RISULTATO D'ESERCIZIO (ante imposte)", f=bold); c2("C15", "=C13", f=bold, nf=EUR, align="right", fill=greyfill)

# ---------------- STATO PATRIMONIALE (right) ----------------
band2("E4:F4", "STATO PATRIMONIALE — a oggi")
c2("E5", "INPUT: Capitale iniziale / PN (cassa apportata)"); c2("F5", 2000, f=bold, fill=inputfill, nf=EUR, align="right")
c2("E6", "ATTIVO", f=bold)
c2("E7", "   Cassa"); c2("F7", f"=F5+{MC}C10-{MC}C7", nf=EUR, align="right")
c2("E8", "   Lavori in corso su ordinazione (WIP / contract asset)"); c2("F8", f"=MAX({MC}C22-{MC}C9,0)", nf=EUR, align="right")
c2("E9", "   Crediti v/clienti"); c2("F9", f"=MAX({MC}C9-{MC}C10,0)", nf=EUR, align="right")
c2("E10", "   TOTALE ATTIVO", f=bold); c2("F10", "=F7+F8+F9", f=bold, nf=EUR, align="right", fill=greyfill)
c2("E11", "PASSIVO E PATRIMONIO NETTO", f=bold)
c2("E12", "   Capitale iniziale"); c2("F12", "=F5", nf=EUR, align="right")
c2("E13", "   Risultato d'esercizio"); c2("F13", "=C15", nf=EUR, align="right")
c2("E14", "   Patrimonio netto", f=bold); c2("F14", "=F12+F13", f=bold, nf=EUR, align="right")
c2("E15", "   Fondo rischi e oneri (perdita su commessa)"); c2("F15", "=C11", nf=EUR, align="right")
c2("E16", "   Anticipi da clienti (contract liability)"); c2("F16", f"=MAX({MC}C9-{MC}C22,0)+MAX({MC}C10-{MC}C9,0)", nf=EUR, align="right")
c2("E17", "   TOTALE PASSIVO E PN", f=bold); c2("F17", "=F14+F15+F16", f=bold, nf=EUR, align="right", fill=greyfill)
c2("E18", "   Quadratura (Attivo − Passivo)", f=bold); c2("F18", "=F10-F17", f=bold, nf=EUR, align="right")

# quadratura conditional format
ws2.conditional_formatting.add("F18", FormulaRule(formula=["ABS(F18)<0.01"], fill=PatternFill("solid", fgColor="CDECD6"), font=Font(color=GREEN, bold=True)))
ws2.conditional_formatting.add("F18", FormulaRule(formula=["ABS(F18)>=0.01"], fill=PatternFill("solid", fgColor="F8CBCB"), font=Font(color=RED, bold=True)))
# EBIT colour
ws2.conditional_formatting.add("C13", FormulaRule(formula=["C13<0"], font=Font(color=RED, bold=True)))
ws2.conditional_formatting.add("C13", FormulaRule(formula=["C13>=0"], font=Font(color=GREEN, bold=True)))

# notes
nb2 = 20
ws2.merge_cells(f"B{nb2}:F{nb2}"); ws2[f"B{nb2}"] = "Logica e ipotesi (didattiche):"; ws2[f"B{nb2}"].font = bold
notes2 = [
    "• Ricavi a CE = prezzo × % completamento (POC). La differenza tra ricavi maturati e fatturato alimenta i Lavori in corso (attivo) o gli Anticipi (passivo).",
    "• Se la commessa è in perdita (EAC > prezzo), l'intera perdita attesa è accantonata subito (Fondo rischi) → vedi voce CE 'Accantonamento'.",
    "• Lo Stato Patrimoniale QUADRA sempre (cella F18 = 0): è costruito in doppia partita a partire dalla commessa.",
    "• Ipotesi semplificatrici: costi tutti pagati per cassa, nessun debito v/fornitori, nessuna imposta/onere finanziario. Scopo solo didattico.",
    "• Prova: nel foglio 'Margine Commessa' porta i Costi a finire (C8) a 6.300 → la commessa va in perdita → EBIT negativo e nasce il Fondo rischi.",
]
for i, t in enumerate(notes2):
    ws2.merge_cells(f"B{nb2+1+i}:F{nb2+1+i}"); ws2[f"B{nb2+1+i}"] = t; ws2[f"B{nb2+1+i}"].font = ital

wb.save(OUT)
print("Saved:", OUT)
